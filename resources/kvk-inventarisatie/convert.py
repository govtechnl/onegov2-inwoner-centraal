"""One-off converter: KvK inventarisatie XLSX -> CSV + grouped Markdown.

Re-run only if the source spreadsheet changes. Outputs are committed alongside.
"""
from __future__ import annotations

import csv
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).parent
SRC = HERE.parent / "KvK - Inventarisatie verplichtingen resultaten - Niet compleet - V2024-12-16.xlsx"
CSV_OUT = HERE / "inventarisatie.csv"
MD_OUT = HERE / "inventarisatie.md"

# Column index (1-based in openpyxl) -> output field name.
# Header is on row 8; sub-header (descriptions) on row 9; data from row 10.
COLUMNS: list[tuple[int, str]] = [
    (1, "volgorde"),
    (2, "instantie"),
    (3, "titel"),
    (4, "type"),
    (5, "oorspronkelijke_eigenaar"),
    (6, "nieuwe_houder"),
    (7, "statuswijziging"),
    (8, "hoe"),
    (11, "titel_communicatie"),
    (12, "ontvanger_communicatie"),
    (13, "communicatie_type"),
    (14, "contactvorm"),
    (15, "gevraagde_actie"),
    (16, "aard"),
    (17, "aard_toelichting"),
    (18, "wanneer"),
    (19, "aantal"),
    (20, "doelgroep"),
    (21, "waarom_eerste_3_maanden"),
    (22, "knelpunten"),
    (23, "bron"),
]


def _clean(v: object) -> str:
    if v is None:
        return ""
    s = str(v).replace("\r", " ").replace("\n", " ").strip()
    while "  " in s:
        s = s.replace("  ", " ")
    return s


def main() -> None:
    wb = load_workbook(SRC, data_only=True)
    sheet = wb.active

    rows: list[dict[str, str]] = []
    for raw in sheet.iter_rows(min_row=10, max_row=sheet.max_row, values_only=True):
        rec = {name: _clean(raw[idx - 1]) for idx, name in COLUMNS}
        if not any(rec.values()):
            continue
        # Skip rows that are just a "volgorde" number with no instantie/titel.
        if not rec["instantie"] and not rec["titel"]:
            continue
        rows.append(rec)

    # CSV
    with CSV_OUT.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=[name for _, name in COLUMNS])
        w.writeheader()
        w.writerows(rows)

    # Markdown grouped by instantie
    grouped: "OrderedDict[str, list[dict[str, str]]]" = OrderedDict()
    for r in rows:
        grouped.setdefault(r["instantie"] or "(onbekend)", []).append(r)

    lines: list[str] = []
    lines.append("# KvK inventarisatie verplichtingen — grouped view")
    lines.append("")
    lines.append(
        "Auto-generated from "
        "`KvK - Inventarisatie verplichtingen resultaten - Niet compleet - V2024-12-16.xlsx`. "
        "See [README.md](README.md) for caveats and provenance."
    )
    lines.append("")
    lines.append(f"**{len(rows)} items across {len(grouped)} instanties.**")
    lines.append("")
    lines.append("## Index")
    lines.append("")
    def _anchor(s: str) -> str:
        # Mimic GitHub's heading-anchor rules: lowercase, drop punctuation, spaces -> -.
        out = []
        for ch in s.lower():
            if ch.isalnum() or ch in (" ", "-"):
                out.append(ch)
        return "".join(out).strip().replace(" ", "-")

    for inst in grouped:
        lines.append(f"- [{inst}](#{_anchor(inst)}) ({len(grouped[inst])})")
    lines.append("")

    for inst, items in grouped.items():
        lines.append(f"## {inst}")
        lines.append("")
        for r in items:
            title = r["titel"] or "(zonder titel)"
            type_ = r["type"]
            head = f"### {title}" + (f" — *{type_}*" if type_ else "")
            lines.append(head)
            lines.append("")
            for label, key in [
                ("Oorspronkelijke eigenaar", "oorspronkelijke_eigenaar"),
                ("Nieuwe houder", "nieuwe_houder"),
                ("Statuswijziging", "statuswijziging"),
                ("Hoe (auto/handmatig)", "hoe"),
                ("Titel communicatie", "titel_communicatie"),
                ("Ontvanger communicatie", "ontvanger_communicatie"),
                ("Communicatie-type", "communicatie_type"),
                ("Contactvorm", "contactvorm"),
                ("Gevraagde actie", "gevraagde_actie"),
                ("Aard", "aard"),
                ("Aard toelichting", "aard_toelichting"),
                ("Wanneer (na overlijden)", "wanneer"),
                ("Aantal", "aantal"),
                ("Doelgroep", "doelgroep"),
                ("Waarom binnen 3 maanden", "waarom_eerste_3_maanden"),
                ("Knelpunten", "knelpunten"),
                ("Bron", "bron"),
            ]:
                val = r[key]
                if val and val != "x":
                    lines.append(f"- **{label}:** {val}")
            lines.append("")
        lines.append("")

    MD_OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {CSV_OUT.relative_to(HERE.parent.parent)} ({len(rows)} rows)")
    print(f"Wrote {MD_OUT.relative_to(HERE.parent.parent)}")


if __name__ == "__main__":
    main()
